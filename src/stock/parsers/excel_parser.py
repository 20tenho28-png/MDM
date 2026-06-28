"""Excel (.xlsx) purchase-guide parser using openpyxl."""
from __future__ import annotations

import io

from stock.parsers.base import (
    ResultadoParse,
    construir_linha,
    mapear_cabecalho,
)


def parse(content: bytes) -> ResultadoParse:
    resultado = ResultadoParse()
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - dependency guard
        resultado.avisos.append("openpyxl não está instalado; não é possível ler XLSX.")
        return resultado

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - report any openpyxl failure as a warning
        resultado.avisos.append(f"Não foi possível abrir o ficheiro Excel: {exc}")
        return resultado

    ws = wb.active
    mapa = None
    for fila in ws.iter_rows(values_only=True):
        celulas = ["" if c is None else c for c in fila]
        if not any(str(c).strip() for c in celulas):
            continue
        if mapa is None:
            mapa = mapear_cabecalho([str(c) for c in celulas])
            if mapa is None:
                continue
            continue
        linha, aviso = construir_linha(list(celulas), mapa)
        if linha is not None:
            resultado.linhas.append(linha)
        elif aviso:
            resultado.avisos.append(aviso)

    wb.close()
    if mapa is None:
        resultado.avisos.append(
            "Não foi encontrado cabeçalho com colunas 'referência' e 'quantidade'."
        )
    return resultado
